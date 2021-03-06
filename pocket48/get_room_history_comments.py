# -*- coding:utf-8 -*-
import requests
import json
import urllib3
import time
import logging
from utils import util
from utils import global_config
from log.my_logger import pocket48_logger as logger
import openpyxl
import emoji
from apscheduler.schedulers.background import BackgroundScheduler, BlockingScheduler

scheduler = BackgroundScheduler()
PA = ''

comment_rst = {}
message_rst = {}
fan_rst = {}

urllib3.disable_warnings()


class Fan:
    def __init__(self, user_id, nick_name, level, gender=0, is_vip=False, verification=False):
        self.user_id = user_id
        self.nick_name = nick_name
        self.level = level
        self.gender = gender
        self.is_vip = is_vip
        self.verification = verification

    def __eq__(self, other):
        return self.user_id == other.user_id

    def __hash__(self):
        return hash(self.user_id)

    def __str__(self):
        return 'user_id: {}, nickname: {}, 等级: {}, 性别: {}, 是否为vip: {}, 是否已实名认证: {}'.format(
            self.user_id, self.nick_name, self.level, self.gender, self.is_vip, self.verification
        )


def get_header():
    global PA
    header = {
        'Content-Type': 'application/json;charset=utf-8',
        'User-Agent': 'PocketFans201807/6.0.16 (iPad; iOS 13.5.1; Scale/2.00)',
        'pa': PA,
        'appInfo': json.dumps({
            'vendor': 'apple',
            'deviceId': 0,
            "appVersion": "6.0.16",
            "appBuild": "200701",
            "osVersion": "13.5.1",
            "osType": "ios",
            "deviceName": "unknow",
            "os": "ios"
        }),
        'token': "c8IOwqXffTUK71AZ5n4WXa/R2JIi32pA+qApuKI3AoecR4kyiG27E4lj/IVXZilOP/CyEYMBxss="
    }
    return header


def get_room_history_msg(member_id, room_id, start_time, end_time=0):
    next_start_time = get_room_msg(member_id, room_id, start_time, end_time)
    while next_start_time >= end_time:
        next_start_time = get_room_msg(member_id, room_id, next_start_time, end_time)
    print('done')


def get_room_msg(member_id, room_id, last_time, end_time):
    print('member_id: {}, room id: {}'.format(member_id, room_id))
    time.sleep(0.5)
    url = 'https://pocketapi.48.cn/im/api/v1/chatroom/msg/list/homeowner'
    params = {
        "ownerId": member_id,
        "roomId": room_id,
        "nextTime": last_time
    }
    try:
        r = requests.post(url, data=json.dumps(params), headers=get_header(), verify=False).text
    except Exception as e:
        print(e)

    try:
        rsp_json = json.loads(r)
        msgs = rsp_json['content']['message']
        if not msgs:
            return -1
        parse_room_msg(msgs, end_time)
        next_time = rsp_json['content']['nextTime']
        return next_time
    except Exception as e:
        print(e)
        return -1


def parse_room_msg(msgs, end_time):
    try:
        for msg in msgs:
            extInfo = json.loads(msg['extInfo'])
            msg_id = msg['msgidClient']  # 消息id
            # rst = cursor.execute("""
            #     SELECT * FROM 'room_message' WHERE message_id=?
            # """, msg_id)

            msg_time_0 = msg["msgTime"]
            if msg_time_0 < end_time:
                return

            msg_time = util.convert_timestamp_to_timestr(msg["msgTime"])
            user_id = extInfo['user']['userId']
            user_name = extInfo['user']['nickName']
            room_id = int(extInfo['roomId'])
            # if int(user_id) != member_messages['id']:
            #     continue
            message_rst[room_id] += 1
            # print('extInfo.keys():' + ','.join(extInfo.keys()))

    except Exception as e:
        print(e)


def get_room_history_comments(room_id, start_time, end_time=0):
    next_start_time = get_room_comments(room_id, start_time, end_time)
    print(end_time)
    while next_start_time >= end_time:
        next_start_time = get_room_comments(room_id, next_start_time, end_time)
    print('done')


def get_room_comments(room_id, last_time, end_time):
    time.sleep(0.5)
    url = 'https://pocketapi.48.cn/im/api/v1/chatroom/msg/list/all'
    params = {
        "roomId": room_id,
        "needTop1Msg": False,
        "nextTime": last_time
    }
    try:
        r = requests.post(url, data=json.dumps(params), headers=get_header(), verify=False).text
    except Exception as e:
        print(e)

    try:
        rsp_json = json.loads(r)
        msgs = rsp_json['content']['message']
        if not msgs:
            return -1
        parse_room_comments(msgs, end_time)
        next_time = rsp_json['content']['nextTime']
        return next_time
    except Exception as e:
        print(e)
        return -1


def parse_room_comments(msgs, end_time):
    try:
        for msg in msgs:
            extInfo = json.loads(msg['extInfo'])
            msg_id = msg['msgidClient']  # 消息id

            msg_time_0 = msg["msgTime"]
            if msg_time_0 < end_time:
                return
            if 'roleId' in extInfo['user'].keys():
                if extInfo['user']['roleId'] != 1:
                    continue

            msg_time = util.convert_timestamp_to_timestr(msg["msgTime"])
            if 'userId' in extInfo['user'].keys():
                user_id = int(extInfo['user']['userId'])
            else:
                user_id = 0
            if 'gender' in extInfo['user'].keys():
                gender = int(extInfo['user']['gender'])
            else:
                gender = 0
            user_name = extInfo['user']['nickName']
            if 'level' in extInfo['user'].keys():
                level = int(extInfo['user']['level'])
            else:
                level = 1
            if 'vip' in extInfo['user'].keys():
                is_vip = extInfo['user']['vip']
            else:
                is_vip = False
            if 'Verification' in extInfo['user'].keys():
                verification = extInfo['user']['Verification']
            else:
                verification = False
            room_id = int(extInfo['sourceId'])

            fan = Fan(user_id, user_name, level, gender, is_vip, verification)
            comment_rst[room_id].add(fan)

            if fan not in fan_rst.keys():
                fan_rst[fan] = set()
            fan_rst[fan].add(room_id)

    except Exception as e:
        print(e)


def get_room_list():
    url = 'https://pocketapi.48.cn/im/api/v1/conversation/page'
    params = {
        "targetType": 0
    }
    try:
        r = requests.post(url, data=json.dumps(params), headers=get_header(), verify=False).json()
    except Exception as e:
        print(e)
    room_list = r['content']['conversations']
    rst = []
    for conv in room_list:
        if conv['ownerId'] == 63:
            continue
        room = {
            "name": conv['ownerName'],
            "id": int(conv['ownerId']),
            "room_id": int(conv['targetId'])
        }
        rst.append(room)
    return rst


@scheduler.scheduled_job('cron', minute="*/5")
def update_pa():
    global PA
    print('更新pa值')
    PA = util.generate_pa2('billjyc', '76e4210ac0808020')
    print('pa: {}'.format(PA))


if __name__ == '__main__':
    update_pa()
    scheduler.start()
    room_list = get_room_list()

    start_time = 1598284800000 + 3 * 86400 * 1000
    end_time = 1598198400000 + 3 * 86400 * 1000
    for i in range(27, 28):
        comment_rst = {}
        message_rst = {}
        fan_rst = {}
        line = 2
        line2 = 2
        line3 = 2
        workbook = openpyxl.Workbook()
        worksheet = workbook.create_sheet('detail')
        worksheet2 = workbook.create_sheet('total')
        worksheet3 = workbook.create_sheet('fans_total')
        worksheet.cell(1, 1, '成员姓名')
        worksheet.cell(1, 2, '房间ID')
        worksheet.cell(1, 3, '粉丝ID')
        worksheet.cell(1, 4, '粉丝昵称')
        worksheet.cell(1, 5, '等级')
        worksheet.cell(1, 6, '性别')
        worksheet.cell(1, 7, '是否为vip')
        worksheet.cell(1, 8, '是否已实名认证')

        worksheet2.cell(1, 1, '成员姓名')
        worksheet2.cell(1, 2, '发送消息数')
        worksheet2.cell(1, 3, '留言粉丝数')
        worksheet2.cell(1, 4, '1-2级粉丝数占比')
        worksheet2.cell(1, 5, '3-6级粉丝数占比')
        worksheet2.cell(1, 6, '7-9级粉丝数占比')
        worksheet2.cell(1, 7, '10级以上粉丝数占比')

        worksheet3.cell(1, 1, '粉丝ID')
        worksheet3.cell(1, 2, '粉丝昵称')
        worksheet3.cell(1, 3, '等级')
        worksheet3.cell(1, 4, '留言成员数')

        print('202007{}'.format(i))
        try:
            for room in room_list:
                print(room['name'])
                if room['room_id'] in [67236601]:
                    continue
                room_id = room['room_id']
                comment_rst[room_id] = set()
                message_rst[room_id] = 0
                get_room_history_msg(room['id'], room['room_id'], start_time, end_time)
                get_room_history_comments(room['room_id'], start_time, end_time)

                level1_2 = 0
                level3_6 = 0
                level7_9 = 0
                level10_12 = 0
                for fan in comment_rst[room_id]:
                    try:
                        print(fan)
                        if int(room['id']) == int(fan.user_id):
                            continue
                        worksheet.cell(line, 1, room['name'])
                        worksheet.cell(line, 2, room_id)
                        worksheet.cell(line, 3, fan.user_id)
                        worksheet.cell(line, 4, emoji.demojize(fan.nick_name))
                        worksheet.cell(line, 5, fan.level)
                        if fan.level <= 2:
                            level1_2 += 1
                        elif fan.level <= 6:
                            level3_6 += 1
                        elif fan.level <= 9:
                            level7_9 += 1
                        else:
                            level10_12 += 1
                        if fan.gender == 0:
                            worksheet.cell(line, 6, '未知')
                        elif fan.gender == 1:
                            worksheet.cell(line, 6, '男')
                        else:
                            worksheet.cell(line, 6, '女')
                        if fan.is_vip:
                            worksheet.cell(line, 7, '是')
                        else:
                            worksheet.cell(line, 7, '否')
                        worksheet.cell(line, 8, '是' if fan.verification else '否')
                        line += 1
                    except Exception as e:
                        logger.exception(e)
                        continue
                worksheet2.cell(line2, 1, room['name'])
                worksheet2.cell(line2, 2, message_rst[room['room_id']])
                print('发送消息数: {}'.format(message_rst[room['room_id']]))
                total = len(comment_rst[room['room_id']])
                worksheet2.cell(line2, 3, total)
                print('留言粉丝数: {}'.format(len(comment_rst[room['room_id']])))
                if total > 0:
                    worksheet2.cell(line2, 4, '%.2f%%' % (level1_2 / total * 100))
                    worksheet2.cell(line2, 5, '%.2f%%' % (level3_6 / total * 100))
                    worksheet2.cell(line2, 6, '%.2f%%' % (level7_9 / total * 100))
                    worksheet2.cell(line2, 7, '%.2f%%' % (level10_12 / total * 100))
                    print('1-2级粉丝占比: {}, 3-6级占比: {}, 7-9级占比: {}, 10级以上占比: {}'.format('%.2f%%' % (level1_2 / total * 100),
                                                                                     '%.2f%%' % (level3_6 / total * 100),
                                                                                     '%.2f%%' % (level7_9 / total * 100),
                                                                                     '%.2f%%' % (level10_12 / total * 100)))
                line2 += 1

            for fan in fan_rst:
                worksheet3.cell(line3, 1, fan.user_id)
                worksheet3.cell(line3, 2, emoji.demojize(fan.nick_name))
                worksheet3.cell(line3, 3, fan.level)
                worksheet3.cell(line3, 4, len(fan_rst[fan]))
                line3 += 1
        except Exception as e:
            print(e)
        finally:
            end_time = start_time
            start_time = start_time + 86400 * 1000
            workbook.save('comments_data_202008{}.xlsx'.format(i))
